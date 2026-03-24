import os
import io
from datetime import datetime, timezone, timedelta, date
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from auth import get_current_user
import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/reports", tags=["reports"])
VM_URL = os.getenv("VICTORIAMETRICS_URL", "http://victoriametrics:8428")
TZ_KST = timezone(timedelta(hours=9))


async def _query_range(query: str, start: int, end: int, step: int = 3600) -> list:
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.get(
            f"{VM_URL}/api/v1/query_range",
            params={"query": query, "start": start, "end": end, "step": step},
        )
        resp.raise_for_status()
        return resp.json().get("data", {}).get("result", [])


def _ts_to_date(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=TZ_KST).strftime("%Y-%m-%d")


def _aggregate(results: list) -> dict:
    """Returns {server_name: {date: [values]}}"""
    data = {}
    for r in results:
        server = r["metric"].get("server_name", "unknown")
        if server not in data:
            data[server] = {}
        for ts, val in r["values"]:
            try:
                v = float(val)
                if v < 0:
                    continue
            except (ValueError, TypeError):
                continue
            data[server].setdefault(_ts_to_date(float(ts)), []).append(v)
    return data


def _daily_stats(data: dict) -> dict:
    """Returns {server_name: {date: {"avg": x, "max": x}}}"""
    result = {}
    for server, dates in data.items():
        result[server] = {}
        for d, vals in dates.items():
            if vals:
                result[server][d] = {
                    "avg": round(sum(vals) / len(vals), 2),
                    "max": round(max(vals), 2),
                }
    return result


@router.get("/range")
async def range_report(
    customer_id: str = Query(...),
    from_date: str = Query(..., description="YYYY-MM-DD"),
    to_date: str = Query(..., description="YYYY-MM-DD"),
    user: dict = Depends(get_current_user),
):
    try:
        start_dt = datetime.strptime(from_date, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, tzinfo=TZ_KST
        )
        end_dt = datetime.strptime(to_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=TZ_KST
        )
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start_dt > end_dt:
        raise HTTPException(status_code=400, detail="from_date must be before to_date")

    days = (end_dt.date() - start_dt.date()).days + 1
    if days > 366:
        raise HTTPException(status_code=400, detail="Range cannot exceed 366 days")

    start_ts = int(start_dt.timestamp())
    end_ts = int(end_dt.timestamp())
    cid = customer_id

    try:
        results = []
        for q in [
            f'100 - avg by(server_name)(rate(node_cpu_seconds_total{{mode="idle",customer_id="{cid}"}}[5m])) * 100',
            f'(1 - node_memory_MemAvailable_bytes{{customer_id="{cid}"}} / node_memory_MemTotal_bytes{{customer_id="{cid}"}}) * 100',
            f'(1 - node_filesystem_avail_bytes{{fstype!~"tmpfs|devtmpfs|overlay|squashfs",customer_id="{cid}",mountpoint="/"}} / node_filesystem_size_bytes{{fstype!~"tmpfs|devtmpfs|overlay|squashfs",customer_id="{cid}",mountpoint="/"}}) * 100',
            f'sum by(server_name)(rate(node_network_receive_bytes_total{{customer_id="{cid}",device!~"lo|docker.*|veth.*|br.*"}}[5m])) / 1048576',
            f'sum by(server_name)(rate(node_network_transmit_bytes_total{{customer_id="{cid}",device!~"lo|docker.*|veth.*|br.*"}}[5m])) / 1048576',
            f'sum by(server_name)(rate(node_disk_read_bytes_total{{customer_id="{cid}"}}[5m])) / 1048576',
            f'sum by(server_name)(rate(node_disk_written_bytes_total{{customer_id="{cid}"}}[5m])) / 1048576',
        ]:
            results.append(await _query_range(q, start_ts, end_ts))
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"VictoriaMetrics query failed: {e}")

    cpu_r, mem_r, disk_r, net_in_r, net_out_r, dr_r, dw_r = results
    cpu = _daily_stats(_aggregate(cpu_r))
    mem = _daily_stats(_aggregate(mem_r))
    disk = _daily_stats(_aggregate(disk_r))
    net_in = _daily_stats(_aggregate(net_in_r))
    net_out = _daily_stats(_aggregate(net_out_r))
    disk_rd = _daily_stats(_aggregate(dr_r))
    disk_wr = _daily_stats(_aggregate(dw_r))

    all_servers = sorted(
        set(cpu) | set(mem) | set(disk) | set(net_in) | set(net_out) | set(disk_rd) | set(disk_wr)
    )

    # All dates in range
    all_dates = []
    cur = start_dt.date()
    while cur <= end_dt.date():
        all_dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{from_date}~{to_date}"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    center = Alignment(horizontal="center")

    headers = [
        "고객사", "서버명", "날짜",
        "CPU 평균(%)", "CPU 최대(%)",
        "메모리 평균(%)", "메모리 최대(%)",
        "디스크 사용률(%)",
        "네트워크 수신(MB/s)", "네트워크 송신(MB/s)",
        "디스크 읽기(MB/s)", "디스크 쓰기(MB/s)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    def g(d, server, dt, stat):
        return d.get(server, {}).get(dt, {}).get(stat, "")

    for server in all_servers:
        for dt in all_dates:
            ws.append([
                customer_id, server, dt,
                g(cpu, server, dt, "avg"), g(cpu, server, dt, "max"),
                g(mem, server, dt, "avg"), g(mem, server, dt, "max"),
                g(disk, server, dt, "avg"),
                g(net_in, server, dt, "avg"), g(net_out, server, dt, "avg"),
                g(disk_rd, server, dt, "avg"), g(disk_wr, server, dt, "avg"),
            ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"report_{customer_id}_{from_date}_{to_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
